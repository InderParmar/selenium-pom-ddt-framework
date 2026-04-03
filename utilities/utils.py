import csv
import json
import logging
import inspect
import os
from datetime import datetime

import softest
from openpyxl import load_workbook

from config.config import LOGS_DIR


class Utils(softest.TestCase):
    """
    Shared utility methods: logging setup, data readers, and soft assertions.
    """

    # ── Logger Factory ───────────────────────────────────────────────────────

    @staticmethod
    def custom_logger(loglevel: int = logging.DEBUG) -> logging.Logger:
        """
        Creates and returns a logger named after the calling class/method.
        Outputs to both console and a timestamped log file under /logs.

        :param loglevel: Minimum log level (default: DEBUG)
        :return:         Configured Logger instance
        """
        caller_name = inspect.stack()[1][3]
        logger = logging.getLogger(caller_name)

        if logger.handlers:
            return logger

        logger.setLevel(loglevel)

        console_handler = logging.StreamHandler()
        console_handler.setLevel(loglevel)
        console_formatter = logging.Formatter(
            fmt="%(asctime)s  [%(levelname)-8s]  %(name)s  →  %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        console_handler.setFormatter(console_formatter)

        log_filename = datetime.now().strftime("test_run_%Y%m%d.log")
        log_filepath = os.path.join(LOGS_DIR, log_filename)

        file_handler = logging.FileHandler(log_filepath, mode="a", encoding="utf-8")
        file_handler.setLevel(logging.DEBUG)
        file_formatter = logging.Formatter(
            fmt="%(asctime)s  [%(levelname)-8s]  %(name)s  →  %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )
        file_handler.setFormatter(file_formatter)

        logger.addHandler(console_handler)
        logger.addHandler(file_handler)
        return logger

    # ── Data Readers ─────────────────────────────────────────────────────────

    @staticmethod
    def read_data_from_json(file_path: str) -> list:
        """
        Reads test data from a JSON array file.
        Each object becomes one test row (as a list of values).

        :param file_path: Path to the .json file
        :return:          List of rows; each row is a list of values
        """
        with open(file_path, mode="r", encoding="utf-8") as f:
            raw = json.load(f)
        data_list = [list(row.values()) for row in raw]
        return data_list

    @staticmethod
    def read_data_from_csv(file_path: str) -> list:
        """
        Reads test data rows from a CSV file (skips header row).

        :param file_path: Path to the CSV file
        :return:          List of rows; each row is a list of cell values
        """
        data_list = []
        with open(file_path, mode="r", encoding="utf-8-sig") as csv_file:
            reader = csv.reader(csv_file)
            next(reader)
            for row in reader:
                data_list.append(row)
        return data_list

    @staticmethod
    def read_data_from_excel(file_path: str, sheet_name: str) -> list:
        """
        Reads test data rows from a specific sheet in an Excel workbook.

        :param file_path:   Path to the .xlsx file
        :param sheet_name:  Name of the worksheet to read
        :return:            List of rows; each row is a list of cell values
        """
        data_list = []
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            data_list.append(list(row))
        return data_list

    # ── Assertions ───────────────────────────────────────────────────────────

    def assert_list_item_text(self, elements: list, expected_text: str):
        """
        Soft-asserts every element before raising any failure.
        Falls back to aria-label attribute if .text is empty.
        Iterates ALL elements first, then raises all failures together at the end.

        :param elements:      List of WebElements
        :param expected_text: Expected text value for every element
        """
        log = logging.getLogger(__name__)
        log.info(
            "Asserting %d result(s) all show '%s'", len(elements), expected_text
        )
        pass_count = 0
        fail_count = 0

        for index, element in enumerate(elements, start=1):
            # Fall back to aria-label when inner text is empty
            actual_text = element.text.strip()
            if not actual_text:
                actual_text = (element.get_attribute("aria-label") or "").strip()
                log.debug(
                    "Element [%d] — .text empty, using aria-label: '%s'",
                    index, actual_text,
                )

            # soft_assert collects failure but does NOT stop the loop
            self.soft_assert(self.assertEqual, actual_text, expected_text)

            if actual_text == expected_text:
                pass_count += 1
                log.debug(
                    "  ✔  [%d/%d] '%s' matched expected '%s'",
                    index, len(elements), actual_text, expected_text,
                )
            else:
                fail_count += 1
                log.warning(
                    "  ✘  [%d/%d] Expected '%s' but got '%s'",
                    index, len(elements), expected_text, actual_text,
                )

        log.info(
            "Assertion summary — PASS: %d | FAIL: %d", pass_count, fail_count
        )

        # Raises here ONLY — after every element has been checked
        self.assert_all()